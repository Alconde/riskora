import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import FileResponse


HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
BODY_ALIGN = Alignment(vertical='top', wrap_text=True)


def exportar_excel(queryset, columns, title, filename, subtitle=None):
    """
    Genera un archivo Excel a partir de un queryset y una lista de columnas.

    columns: lista de dicts con keys:
        - 'header': str (nombre de la columna)
        - 'value': callable(row) -> value
        - 'width': int (ancho opcional, default 25)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    num_cols = len(columns)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    ws['A1'] = title
    ws['A1'].font = Font(name='Arial', bold=True, size=14)

    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
        ws['A2'] = subtitle
        ws['A2'].font = Font(name='Arial', size=10, color='64748B')
        header_row = 4
    else:
        header_row = 3

    ws.row_dimensions[header_row - 1].height = 6

    for col_num, col in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_num, value=col['header'])
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    for row_num, obj in enumerate(queryset, header_row + 1):
        for col_num, col in enumerate(columns, 1):
            value = col['value'](obj)
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = THIN_BORDER
            cell.alignment = BODY_ALIGN

    for i, col in enumerate(columns, 1):
        width = col.get('width', 25)
        col_letter = chr(64 + i) if i <= 26 else chr(64 + (i - 1) // 26) + chr(65 + (i - 1) % 26)
        ws.column_dimensions[col_letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return FileResponse(
        buffer,
        as_attachment=True,
        filename=filename,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
