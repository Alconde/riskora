from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import JsonResponse, HttpResponse, FileResponse
from django.urls import reverse_lazy
from django.views.generic import (
    TemplateView, ListView, DetailView, CreateView, UpdateView, DeleteView,
)
import io

from apps.risk_assessment.forms import (
    EvaluacionRiesgosForm, ItemEvaluacionRiesgosForm, InformeRiesgoEspecialForm,
)
from apps.risk_assessment.models import (
    EvaluacionRiesgos, ItemEvaluacionRiesgos, TipoPeligro, InformeRiesgoEspecial,
)
from apps.risk_assessment.services import calcular_grado_riesgo, calcular_estadisticas_evaluacion
from apps.companies.models import Company
from apps.core.mixins import CompanyScopedMixin


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

class EvaluacionDashboardView(LoginRequiredMixin, CompanyScopedMixin, TemplateView):
    template_name = 'risk_assessment/dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        empresa = self.get_active_company()
        if empresa:
            items = ItemEvaluacionRiesgos.objects.filter(evaluacion__empresa=empresa)
            ctx['total_evitables'] = items.filter(tipo_riesgo='evitable').count()
            ctx['total_monitorizables'] = items.filter(tipo_riesgo='monitorizable').count()
            ctx['total_no_evitables'] = items.filter(tipo_riesgo='no_evitable').count()
            ctx['total_evaluaciones'] = EvaluacionRiesgos.objects.filter(empresa=empresa).count()
            ctx['informes_higienico'] = InformeRiesgoEspecial.objects.filter(company=empresa, tipo='higienico').count()
            ctx['informes_psicosocial'] = InformeRiesgoEspecial.objects.filter(company=empresa, tipo='psicosocial').count()
            ctx['informes_ergonomico'] = InformeRiesgoEspecial.objects.filter(company=empresa, tipo='ergonomico').count()
        return ctx


# ---------------------------------------------------------------------------
# Listas filtradas por tipo de riesgo
# ---------------------------------------------------------------------------

class RiesgosPorTipoListView(LoginRequiredMixin, CompanyScopedMixin, ListView):
    model = ItemEvaluacionRiesgos
    template_name = 'risk_assessment/riesgos_lista.html'
    context_object_name = 'items'
    paginate_by = 30

    tipo_riesgo = 'evitable'

    def get_queryset(self):
        qs = super().get_queryset()
        empresa = self.get_active_company()
        if empresa:
            qs = qs.filter(evaluacion__empresa=empresa, tipo_riesgo=self.tipo_riesgo)
        else:
            qs = qs.none()
        return qs.select_related('evaluacion', 'puesto_trabajo', 'tipo_peligro', 'responsable_medida')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['tipo_riesgo'] = self.tipo_riesgo
        ctx['tipo_label'] = dict(ItemEvaluacionRiesgos.TipoRiesgo.choices).get(self.tipo_riesgo, '')
        empresa = self.get_active_company()
        if empresa:
            all_items = ItemEvaluacionRiesgos.objects.filter(evaluacion__empresa=empresa)
            ctx['total_evitables'] = all_items.filter(tipo_riesgo='evitable').count()
            ctx['total_monitorizables'] = all_items.filter(tipo_riesgo='monitorizable').count()
            ctx['total_no_evitables'] = all_items.filter(tipo_riesgo='no_evitable').count()
        return ctx


class RiesgosEvitablesView(RiesgosPorTipoListView):
    tipo_riesgo = 'evitable'


class RiesgosMonitorizablesView(RiesgosPorTipoListView):
    tipo_riesgo = 'monitorizable'


class RiesgosNoEvitablesView(RiesgosPorTipoListView):
    tipo_riesgo = 'no_evitable'


# ---------------------------------------------------------------------------
# Informes especiales (higiénico, psicosocial, ergonómico)
# ---------------------------------------------------------------------------

class InformesEspecialesListView(LoginRequiredMixin, CompanyScopedMixin, ListView):
    model = InformeRiesgoEspecial
    template_name = 'risk_assessment/informes_especiales.html'
    context_object_name = 'informes'
    paginate_by = 20

    tipo_informe = 'higienico'

    def get_queryset(self):
        qs = super().get_queryset()
        empresa = self.get_active_company()
        if empresa:
            qs = qs.filter(company=empresa, tipo=self.tipo_informe)
        else:
            qs = qs.none()
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['tipo_informe'] = self.tipo_informe
        ctx['tipo_label'] = dict(InformeRiesgoEspecial.TIPO_CHOICES).get(self.tipo_informe, '')
        return ctx


class InformesHigienicoView(InformesEspecialesListView):
    tipo_informe = 'higienico'


class InformesPsicosocialView(InformesEspecialesListView):
    tipo_informe = 'psicosocial'


class InformesErgonomicoView(InformesEspecialesListView):
    tipo_informe = 'ergonomico'


class InformeEspecialCreateView(LoginRequiredMixin, CompanyScopedMixin, CreateView):
    model = InformeRiesgoEspecial
    form_class = InformeRiesgoEspecialForm
    template_name = 'risk_assessment/informe_especial_form.html'

    def get_success_url(self):
        tipo = self.object.tipo
        url_map = {
            'higienico': 'risk_assessment:informes-higienico',
            'psicosocial': 'risk_assessment:informes-psicosocial',
            'ergonomico': 'risk_assessment:informes-ergonomico',
        }
        return reverse_lazy(url_map.get(tipo, 'risk_assessment:dashboard'))

    def form_valid(self, form):
        form.instance.company = self.get_active_company()
        return super().form_valid(form)


class InformeEspecialDeleteView(LoginRequiredMixin, CompanyScopedMixin, DeleteView):
    model = InformeRiesgoEspecial
    template_name = 'risk_assessment/confirm_delete_simple.html'

    def get_success_url(self):
        tipo = self.object.tipo
        url_map = {
            'higienico': 'risk_assessment:informes-higienico',
            'psicosocial': 'risk_assessment:informes-psicosocial',
            'ergonomico': 'risk_assessment:informes-ergonomico',
        }
        return reverse_lazy(url_map.get(tipo, 'risk_assessment:dashboard'))


# ---------------------------------------------------------------------------
# Evaluaciones (CRUD original)
# ---------------------------------------------------------------------------

class EvaluacionRiesgosListView(LoginRequiredMixin, CompanyScopedMixin, ListView):
    model = EvaluacionRiesgos
    template_name = 'risk_assessment/evaluacion_list.html'
    context_object_name = 'evaluaciones'
    paginate_by = 20
    login_url = '/login/'
    company_field_name = 'empresa'

    def get_base_queryset(self):
        return EvaluacionRiesgos.objects.select_related(
            'empresa', 'centro_trabajo', 'revisado_por'
        )

    def get_queryset(self):
        queryset = self.get_company_scoped_queryset(self.get_base_queryset())

        q = self.request.GET.get('q', '').strip()
        estado = self.request.GET.get('estado', '').strip()

        if q:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(titulo__icontains=q) |
                Q(centro_trabajo__name__icontains=q)
            )

        if estado:
            queryset = queryset.filter(estado=estado)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filtros'] = self.request.GET
        context['estado_choices'] = EvaluacionRiesgos.Estado.choices
        return context


class EvaluacionRiesgosDetailView(LoginRequiredMixin, CompanyScopedMixin, DetailView):
    model = EvaluacionRiesgos
    template_name = 'risk_assessment/evaluacion_detail.html'
    context_object_name = 'evaluacion'
    login_url = '/login/'
    company_field_name = 'empresa'

    def get_queryset(self):
        return EvaluacionRiesgos.objects.select_related(
            'empresa', 'centro_trabajo', 'revisado_por', 'aprobado_por'
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        items = self.object.items.select_related(
            'puesto_trabajo', 'tipo_peligro', 'responsable_medida'
        )
        context['items'] = items
        context['stats'] = calcular_estadisticas_evaluacion(items)
        context['item_form'] = ItemEvaluacionRiesgosForm(
            empresa=self.object.empresa
        )
        context['items_evitables'] = items.filter(tipo_riesgo='evitable')
        context['items_monitorizables'] = items.filter(tipo_riesgo='monitorizable')
        context['items_no_evitables'] = items.filter(tipo_riesgo='no_evitable')
        return context


class EvaluacionRiesgosCreateView(LoginRequiredMixin, CompanyScopedMixin, CreateView):
    model = EvaluacionRiesgos
    form_class = EvaluacionRiesgosForm
    template_name = 'risk_assessment/evaluacion_form.html'
    success_url = reverse_lazy('risk_assessment:evaluacion-list')
    login_url = '/login/'
    company_field_name = 'empresa'

    def form_valid(self, form):
        active_company = self.get_active_company()
        if active_company:
            form.instance.empresa = active_company
        return super().form_valid(form)


class EvaluacionRiesgosUpdateView(LoginRequiredMixin, CompanyScopedMixin, UpdateView):
    model = EvaluacionRiesgos
    form_class = EvaluacionRiesgosForm
    template_name = 'risk_assessment/evaluacion_form.html'
    login_url = '/login/'
    company_field_name = 'empresa'

    def get_queryset(self):
        return self.get_company_scoped_queryset(
            EvaluacionRiesgos.objects.select_related(
                'empresa', 'centro_trabajo', 'revisado_por', 'aprobado_por'
            )
        )

    def get_success_url(self):
        return reverse_lazy('risk_assessment:evaluacion-detail', kwargs={'pk': self.object.pk})


class EvaluacionRiesgosDeleteView(LoginRequiredMixin, CompanyScopedMixin, DeleteView):
    model = EvaluacionRiesgos
    template_name = 'risk_assessment/evaluacion_confirm_delete.html'
    success_url = reverse_lazy('risk_assessment:evaluacion-list')
    login_url = '/login/'
    company_field_name = 'empresa'

    def get_queryset(self):
        return self.get_company_scoped_queryset(
            EvaluacionRiesgos.objects.select_related('empresa')
        )


class ItemEvaluacionCreateView(LoginRequiredMixin, CreateView):
    model = ItemEvaluacionRiesgos
    form_class = ItemEvaluacionRiesgosForm
    template_name = 'risk_assessment/item_form.html'
    login_url = '/login/'

    def get_evaluacion(self):
        return EvaluacionRiesgos.objects.get(pk=self.kwargs['evaluacion_pk'])

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        evaluacion = self.get_evaluacion()
        kwargs['empresa'] = evaluacion.empresa
        return kwargs

    def form_valid(self, form):
        evaluacion = self.get_evaluacion()
        form.instance.evaluacion = evaluacion
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        empresa = self.get_evaluacion().empresa
        from django.db.models import Q
        from apps.preventive_planning.models import MedidaPreventivaCatalogo
        qs = MedidaPreventivaCatalogo.objects.filter(
            Q(company__isnull=True) | Q(company=empresa),
            activo=True,
        )
        context['medidas_catalogo_agrupadas'] = [
            (cat_label, qs.filter(categoria=cat_val))
            for cat_val, cat_label in MedidaPreventivaCatalogo.Categoria.choices
            if qs.filter(categoria=cat_val).exists()
        ]
        context['medidas_seleccionadas'] = set()
        return context

    def get_success_url(self):
        return reverse_lazy(
            'risk_assessment:evaluacion-detail', kwargs={'pk': self.kwargs['evaluacion_pk']}
        )


class ItemEvaluacionUpdateView(LoginRequiredMixin, UpdateView):
    model = ItemEvaluacionRiesgos
    form_class = ItemEvaluacionRiesgosForm
    template_name = 'risk_assessment/item_form.html'
    login_url = '/login/'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['empresa'] = self.object.evaluacion.empresa
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        empresa = self.object.evaluacion.empresa
        from django.db.models import Q
        from apps.preventive_planning.models import MedidaPreventivaCatalogo
        qs = MedidaPreventivaCatalogo.objects.filter(
            Q(company__isnull=True) | Q(company=empresa),
            activo=True,
        )
        context['medidas_catalogo_agrupadas'] = [
            (cat_label, qs.filter(categoria=cat_val))
            for cat_val, cat_label in MedidaPreventivaCatalogo.Categoria.choices
            if qs.filter(categoria=cat_val).exists()
        ]
        selected = set(self.object.medidas_catalogo.values_list('pk', flat=True))
        if self.request.POST:
            selected = set(self.request.POST.getlist('medidas_catalogo'))
        context['medidas_seleccionadas'] = selected
        return context

    def get_success_url(self):
        return reverse_lazy(
            'risk_assessment:evaluacion-detail',
            kwargs={'pk': self.object.evaluacion.pk},
        )


class ItemEvaluacionDeleteView(LoginRequiredMixin, DeleteView):
    model = ItemEvaluacionRiesgos
    template_name = 'risk_assessment/item_confirm_delete.html'
    login_url = '/login/'

    def get_success_url(self):
        return reverse_lazy(
            'risk_assessment:evaluacion-detail',
            kwargs={'pk': self.object.evaluacion.pk},
        )


def calcular_riesgo_ajax(request):
    """Endpoint AJAX para calcular el grado de riesgo en tiempo real."""
    try:
        probabilidad = int(request.GET.get('probabilidad', 0))
        severidad = int(request.GET.get('severidad', 0))
        resultado = calcular_grado_riesgo(probabilidad, severidad)
        return JsonResponse(resultado)
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Valores inválidos'}, status=400)


# ---------------------------------------------------------------------------
# Importación / Exportación Excel
# ---------------------------------------------------------------------------

ENCabezado_EXCEL = [
    'Puesto de trabajo',
    'Tipo de peligro',
    'Factor de riesgo / condición detectada',
    'Riesgo',
    'Medidas existentes',
    'Probabilidad (1-3)',
    'Severidad (1-3)',
    'Medidas propuestas',
]


def _obtener_opciones_riesgo():
    return dict(ItemEvaluacionRiesgos.RIESGO_CHOICES)


def _obtener_opciones_prob_sev():
    prob = dict(ItemEvaluacionRiesgos.Probabilidad.choices)
    sev = dict(ItemEvaluacionRiesgos.Severidad.choices)
    return prob, sev


def exportar_excel(request, evaluacion_pk):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    evaluacion = EvaluacionRiesgos.objects.get(pk=evaluacion_pk)
    items = evaluacion.items.select_related(
        'puesto_trabajo', 'tipo_peligro'
    ).order_by('puesto_trabajo', 'factor_riesgo_condicion')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Evaluación de Riesgos'

    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    ws.merge_cells('A1:H1')
    ws['A1'] = evaluacion.titulo
    ws['A1'].font = Font(name='Arial', bold=True, size=14)
    ws.merge_cells('A2:H2')
    ws['A2'] = (
        f'Empresa: {evaluacion.empresa} | Centro: {evaluacion.centro_trabajo} | '
        f'Fecha: {evaluacion.fecha_evaluacion.strftime("%d/%m/%Y")}'
    )
    ws['A2'].font = Font(name='Arial', size=10, color='64748B')
    ws.row_dimensions[3].height = 6

    for col_num, header in enumerate(ENCabezado_EXCEL, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    prob_labels, sev_labels = _obtener_opciones_prob_sev()
    riesgo_opciones = _obtener_opciones_riesgo()

    for row_num, item in enumerate(items, 5):
        data = [
            str(item.puesto_trabajo),
            str(item.tipo_peligro) if item.tipo_peligro else '',
            item.factor_riesgo_condicion,
            riesgo_opciones.get(item.riesgo, item.riesgo),
            item.medidas_existentes,
            item.probabilidad,
            item.severidad,
            item.medidas_propuestas,
        ]
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    col_widths = [25, 25, 40, 30, 35, 15, 15, 35]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f'evaluacion_riesgos_{evaluacion.pk}.xlsx'
    response = FileResponse(
        buffer, as_attachment=True, filename=filename,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    return response


def importar_excel(request, evaluacion_pk):
    from openpyxl import load_workbook

    evaluacion = EvaluacionRiesgos.objects.get(pk=evaluacion_pk)

    if request.method != 'POST':
        return HttpResponse('Método no permitido', status=405)

    archivo = request.FILES.get('archivo_excel')
    if not archivo:
        return HttpResponse('No se proporcionó archivo', status=400)

    try:
        wb = load_workbook(archivo, read_only=True)
        ws = wb.active

        from apps.workers.models import JobPosition
        puestos = {p.name: p for p in JobPosition.objects.filter(company=evaluacion.empresa)}
        tipos = {t.nombre: t for t in TipoPeligro.objects.filter(activo=True)}
        riesgo_opciones_inv = {v.lower(): k for k, v in _obtener_opciones_riesgo().items()}

        start_row = 1
        for row in ws.iter_rows(min_row=1, max_row=10, max_col=1):
            for cell in row:
                if cell.value and 'puesto' in str(cell.value).lower():
                    start_row = cell.row + 1
                    break

        importados = 0
        errores = []

        for row_num, row in enumerate(ws.iter_rows(min_row=start_row, max_col=8, values_only=True)):
            if not row[0]:
                continue

            puesto_nombre = str(row[0]).strip()
            tipo_nombre = str(row[1]).strip() if row[1] else ''
            factor_riesgo = str(row[2]).strip() if row[2] else ''
            riesgo_texto = str(row[3]).strip().lower() if row[3] else ''
            medidas_existentes = str(row[4]).strip() if row[4] else ''
            probabilidad = row[5]
            severidad = row[6]
            medidas_propuestas = str(row[7]).strip() if row[7] else ''

            if not factor_riesgo:
                errores.append(f'Fila {row_num + start_row}: falta factor de riesgo')
                continue
            if not probabilidad or not severidad:
                errores.append(f'Fila {row_num + start_row}: falta probabilidad o severidad')
                continue
            try:
                probabilidad = int(probabilidad)
                severidad = int(severidad)
            except (ValueError, TypeError):
                errores.append(f'Fila {row_num + start_row}: probabilidad/severidad no son números')
                continue
            if probabilidad not in (1, 2, 3) or severidad not in (1, 2, 3):
                errores.append(f'Fila {row_num + start_row}: probabilidad/severidad fuera de rango (1-3)')
                continue

            puesto = puestos.get(puesto_nombre)
            if not puesto:
                errores.append(f'Fila {row_num + start_row}: puesto "{puesto_nombre}" no encontrado')
                continue

            tipo_peligro = tipos.get(tipo_nombre) if tipo_nombre else None
            riesgo_valor = riesgo_opciones_inv.get(riesgo_texto, '')
            resultado = calcular_grado_riesgo(probabilidad, severidad)

            ItemEvaluacionRiesgos.objects.create(
                evaluacion=evaluacion,
                puesto_trabajo=puesto,
                tipo_peligro=tipo_peligro,
                factor_riesgo_condicion=factor_riesgo,
                riesgo=riesgo_valor,
                medidas_existentes=medidas_existentes,
                probabilidad=probabilidad,
                severidad=severidad,
                grado_riesgo=resultado['grado'],
                nivel_riesgo=resultado['nivel'],
                medidas_propuestas=medidas_propuestas,
            )
            importados += 1

        wb.close()

        return HttpResponse(
            f'<html><body>'
            f'<h2>Importación completada</h2>'
            f'<p>Items importados: <strong>{importados}</strong></p>'
            f'{"<p>Errores:</p><ul>" + "".join("<li>" + e + "</li>" for e in errores) + "</ul>" if errores else ""}'
            f'<p><a href="/evaluaciones/{evaluacion.pk}/">Volver a la evaluación</a></p>'
            f'</body></html>'
        )

    except Exception as e:
        return HttpResponse(f'Error al procesar el archivo: {str(e)}', status=400)


def plantilla_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = 'Plantilla Evaluación'

    ws.merge_cells('A1:H1')
    ws['A1'] = 'Plantilla de Importación - Evaluación de Riesgos'
    ws['A1'].font = Font(name='Arial', bold=True, size=14)
    ws.merge_cells('A2:H2')
    ws['A2'] = 'Rellena esta plantilla y sube el archivo en la evaluación correspondiente.'
    ws['A2'].font = Font(name='Arial', size=10, color='64748B')
    ws.merge_cells('A3:H3')
    ws['A3'] = (
        'Probabilidad: 1=Baja, 2=Media, 3=Alta  |  '
        'Severidad: 1=Baja, 2=Media, 3=Alta'
    )
    ws['A3'].font = Font(name='Arial', size=9, italic=True, color='64748B')
    ws.row_dimensions[4].height = 6

    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    for col_num, header in enumerate(ENCabezado_EXCEL, 1):
        cell = ws.cell(row=5, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ejemplo = [
        'Soldador', 'Ruido', 'Ruido continuo en zona de soldadura',
        'Sordera', 'Tapones auditivos', 3, 2,
        'Instalar cabinas de soldadura con insonorización',
    ]
    for col_num, value in enumerate(ejemplo, 1):
        cell = ws.cell(row=6, column=col_num, value=value)
        cell.border = thin_border
        cell.font = Font(name='Arial', size=10, color='94A3B8')

    col_widths = [25, 25, 40, 30, 35, 15, 15, 35]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = FileResponse(
        buffer, as_attachment=True, filename='plantilla_evaluacion_riesgos.xlsx',
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    return response
