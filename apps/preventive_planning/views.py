from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView
from django.http import HttpResponse
from django.db.models import Q

from apps.core.mixins import CompanyScopedMixin
from .models import MedidaPreventivaCatalogo, ItemPlanificacion
from .forms import MedidaPreventivaCatalogoForm, ItemPlanificacionForm
from .services import calcular_estadisticas_planificacion, importar_planificacion_excel


class PlanificacionDashboardView(LoginRequiredMixin, CompanyScopedMixin, TemplateView):
    template_name = 'preventive_planning/dashboard.html'
    login_url = '/login/'
    company_field_name = 'empresa'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        empresa = self.get_active_company()
        context['stats'] = calcular_estadisticas_planificacion(empresa)
        if empresa:
            context['ultimos_items'] = ItemPlanificacion.objects.filter(
                empresa=empresa
            ).order_by('-created_at')[:10]
        else:
            context['ultimos_items'] = []
        return context


class ItemPlanificacionListView(LoginRequiredMixin, CompanyScopedMixin, ListView):
    model = ItemPlanificacion
    template_name = 'preventive_planning/item_list.html'
    context_object_name = 'items'
    paginate_by = 20
    login_url = '/login/'
    company_field_name = 'empresa'

    def get_queryset(self):
        qs = super().get_queryset()
        empresa = self.get_active_company()
        if empresa:
            qs = qs.filter(empresa=empresa)

        q = self.request.GET.get('q', '').strip()
        if q:
            qs = qs.filter(
                Q(factor_riesgo__icontains=q) |
                Q(ambito_puesto__icontains=q) |
                Q(responsable__icontains=q)
            )

        estado = self.request.GET.get('estado', '')
        if estado:
            qs = qs.filter(estado=estado)

        tipo = self.request.GET.get('tipo', '')
        if tipo:
            qs = qs.filter(tipo_factor_riesgo=tipo)

        return qs


class ItemPlanificacionDetailView(LoginRequiredMixin, CompanyScopedMixin, DetailView):
    model = ItemPlanificacion
    template_name = 'preventive_planning/item_detail.html'
    context_object_name = 'item'
    login_url = '/login/'
    company_field_name = 'empresa'


class ItemPlanificacionCreateView(LoginRequiredMixin, CompanyScopedMixin, CreateView):
    model = ItemPlanificacion
    form_class = ItemPlanificacionForm
    template_name = 'preventive_planning/item_form.html'
    login_url = '/login/'
    company_field_name = 'empresa'

    def form_valid(self, form):
        empresa = self.get_active_company()
        if empresa:
            form.instance.empresa = empresa
        return super().form_valid(form)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['empresa'] = self.get_active_company()
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        empresa = self.get_active_company()
        from django.db.models import Q
        qs = MedidaPreventivaCatalogo.objects.filter(
            Q(company__isnull=True) | Q(company=empresa),
            activo=True,
        ) if empresa else MedidaPreventivaCatalogo.objects.filter(
            company__isnull=True, activo=True,
        )
        context['medidas_catalogo_agrupadas'] = [
            (cat_label, qs.filter(categoria=cat_val))
            for cat_val, cat_label in MedidaPreventivaCatalogo.Categoria.choices
            if qs.filter(categoria=cat_val).exists()
        ]
        selected = set()
        if self.object and self.object.pk:
            selected = set(self.object.medidas_catalogo.values_list('pk', flat=True))
        elif self.request.POST:
            selected = set(self.request.POST.getlist('medidas_catalogo'))
        context['medidas_seleccionadas'] = selected
        return context

    def get_success_url(self):
        return reverse_lazy('preventive_planning:item-detail', kwargs={'pk': self.object.pk})


class ItemPlanificacionUpdateView(LoginRequiredMixin, CompanyScopedMixin, UpdateView):
    model = ItemPlanificacion
    form_class = ItemPlanificacionForm
    template_name = 'preventive_planning/item_form.html'
    login_url = '/login/'
    company_field_name = 'empresa'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['empresa'] = self.get_active_company()
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        empresa = self.get_active_company()
        from django.db.models import Q
        qs = MedidaPreventivaCatalogo.objects.filter(
            Q(company__isnull=True) | Q(company=empresa),
            activo=True,
        ) if empresa else MedidaPreventivaCatalogo.objects.filter(
            company__isnull=True, activo=True,
        )
        context['medidas_catalogo_agrupadas'] = [
            (cat_label, qs.filter(categoria=cat_val))
            for cat_val, cat_label in MedidaPreventivaCatalogo.Categoria.choices
            if qs.filter(categoria=cat_val).exists()
        ]
        selected = set()
        if self.object and self.object.pk:
            selected = set(self.object.medidas_catalogo.values_list('pk', flat=True))
        elif self.request.POST:
            selected = set(self.request.POST.getlist('medidas_catalogo'))
        context['medidas_seleccionadas'] = selected
        return context

    def get_success_url(self):
        return reverse_lazy('preventive_planning:item-detail', kwargs={'pk': self.object.pk})


class ItemPlanificacionDeleteView(LoginRequiredMixin, CompanyScopedMixin, DeleteView):
    model = ItemPlanificacion
    template_name = 'preventive_planning/item_confirm_delete.html'
    login_url = '/login/'
    company_field_name = 'empresa'
    success_url = reverse_lazy('preventive_planning:item-list')


def importar_excel(request):
    if request.method != 'POST':
        return HttpResponse('Metodo no permitido', status=405)

    archivo = request.FILES.get('archivo_excel')
    if not archivo:
        return HttpResponse('No se proporciono archivo', status=400)

    empresa = getattr(request, 'active_company', None)
    if not empresa:
        return HttpResponse('No hay empresa activa', status=400)

    importados, errores = importar_planificacion_excel(archivo, empresa)

    return HttpResponse(
        f'<html><body>'
        f'<h2>Importacion completada</h2>'
        f'<p>Items importados: <strong>{importados}</strong></p>'
        f'{"<p>Errores:</p><ul>" + "".join("<li>" + e + "</li>" for e in errores) + "</ul>" if errores else ""}'
        f'<p><a href="{reverse_lazy("preventive_planning:item-list")}">Volver al listado</a></p>'
        f'</body></html>'
    )


def plantilla_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = 'Planificacion Preventiva'

    ws.merge_cells('A1:P1')
    ws['A1'] = 'Plantilla de Importacion - Planificacion de la Actividad Preventiva'
    ws['A1'].font = Font(name='Arial', bold=True, size=14)
    ws.merge_cells('A2:P2')
    ws['A2'] = 'Rellena esta plantilla y sube el archivo en el apartado de Planificacion.'
    ws['A2'].font = Font(name='Arial', size=10, color='64748B')

    headers = [
        'Ambito/Puesto de trabajo', 'Tipo de factor de riesgo',
        'Factor de riesgo', 'Detalle', 'Riesgos', 'PB', 'SV', 'GR',
        'Medidas preventivas', 'Detalle de la medida - Accion correctora',
        'Plazo limite', 'Fecha Objetivo', 'Responsable', 'Costes',
        'Estado', 'Origen',
    ]

    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='0F766E', end_color='0F766E', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ejemplo = [
        'Todos los ambitos', 'Evitables', 'Cuadros electricos sin senal',
        '', 'Contactos electricos', 'M', 'M', 'M',
        'Colocar senal de riesgo electrico', ' senal en panel',
        '6 meses', '2025-03-24', 'Patricia Conde', '1-50',
        'Pendiente', '',
    ]
    for col_num, value in enumerate(ejemplo, 1):
        cell = ws.cell(row=5, column=col_num, value=value)
        cell.border = thin_border
        cell.font = Font(name='Arial', size=10, color='94A3B8')

    col_widths = [25, 20, 40, 30, 35, 8, 8, 8, 40, 40, 12, 15, 20, 15, 12, 15]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_planificacion_preventiva.xlsx"'
    return response


def exportar_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io

    empresa = getattr(request, 'active_company', None)
    if not empresa:
        return HttpResponse('No hay empresa activa', status=400)

    items = ItemPlanificacion.objects.filter(empresa=empresa).order_by(
        'tipo_factor_riesgo', 'factor_riesgo'
    )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Planificacion Preventiva'

    headers = [
        'Ambito/Puesto de trabajo', 'Tipo de factor de riesgo',
        'Factor de riesgo', 'Detalle', 'Riesgos', 'PB', 'SV', 'GR',
        'Medidas preventivas', 'Detalle de la medida',
        'Plazo limite', 'Fecha Objetivo', 'Responsable', 'Costes',
        'Estado', 'Origen',
    ]

    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='0F766E', end_color='0F766E', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    tipo_display = dict(ItemPlanificacion.TipoFactorRiesgo.choices)
    estado_display = dict(ItemPlanificacion.Estado.choices)
    riesgo_display = dict(ItemPlanificacion.RIESGO_CHOICES)

    for row_num, item in enumerate(items, 2):
        data = [
            item.ambito_puesto,
            tipo_display.get(item.tipo_factor_riesgo, ''),
            item.factor_riesgo,
            item.detalle,
            riesgo_display.get(item.riesgos, item.riesgos),
            item.pb,
            item.sv,
            item.gr,
            item.medidas_preventivas,
            item.detalle_medida,
            item.plazo_limite,
            str(item.fecha_objetivo) if item.fecha_objetivo else '',
            item.responsable,
            item.costes,
            estado_display.get(item.estado, ''),
            item.origen,
        ]
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    col_widths = [25, 20, 40, 30, 35, 8, 8, 8, 40, 40, 12, 15, 20, 15, 12, 15]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="planificacion_preventiva_{empresa.pk}.xlsx"'
    return response


class MedidaCatalogoListView(LoginRequiredMixin, CompanyScopedMixin, ListView):
    model = MedidaPreventivaCatalogo
    template_name = 'preventive_planning/catalogo_list.html'
    context_object_name = 'medidas'
    paginate_by = 50
    login_url = '/login/'
    company_field_name = 'empresa'

    def get_queryset(self):
        qs = super().get_queryset()
        empresa = self.get_active_company()
        qs = qs.filter(
            Q(company__isnull=True) | Q(company=empresa),
            activo=True,
        )
        q = self.request.GET.get('q', '').strip()
        if q:
            qs = qs.filter(
                Q(nombre__icontains=q) | Q(normativa__icontains=q)
            )
        cat = self.request.GET.get('categoria', '')
        if cat:
            qs = qs.filter(categoria=cat)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categorias'] = MedidaPreventivaCatalogo.Categoria.choices
        context['categoria_actual'] = self.request.GET.get('categoria', '')
        return context


class MedidaCatalogoCreateView(LoginRequiredMixin, CompanyScopedMixin, CreateView):
    model = MedidaPreventivaCatalogo
    form_class = MedidaPreventivaCatalogoForm
    template_name = 'preventive_planning/catalogo_form.html'
    login_url = '/login/'
    company_field_name = 'empresa'

    def form_valid(self, form):
        empresa = self.get_active_company()
        form.instance.company = empresa
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('preventive_planning:catalogo-list')


class MedidaCatalogoUpdateView(LoginRequiredMixin, CompanyScopedMixin, UpdateView):
    model = MedidaPreventivaCatalogo
    form_class = MedidaPreventivaCatalogoForm
    template_name = 'preventive_planning/catalogo_form.html'
    login_url = '/login/'
    company_field_name = 'empresa'

    def get_success_url(self):
        return reverse_lazy('preventive_planning:catalogo-list')


class MedidaCatalogoDeleteView(LoginRequiredMixin, CompanyScopedMixin, DeleteView):
    model = MedidaPreventivaCatalogo
    template_name = 'preventive_planning/catalogo_confirm_delete.html'
    login_url = '/login/'
    company_field_name = 'empresa'
    success_url = reverse_lazy('preventive_planning:catalogo-list')
