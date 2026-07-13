from django.db import models


def calcular_estadisticas_planificacion(empresa=None):
    from .models import ItemPlanificacion

    qs = ItemPlanificacion.objects.all()
    if empresa:
        qs = qs.filter(empresa=empresa)

    total = qs.count()
    pendientes = qs.filter(estado='pendiente').count()
    en_curso = qs.filter(estado='en_curso').count()
    implementadas = qs.filter(estado='implementada').count()
    continuas = qs.filter(estado='continua').count()
    evitables = qs.filter(tipo_factor_riesgo='evitables').count()
    monitorizables = qs.filter(tipo_factor_riesgo='monitorizables').count()

    return {
        'total': total,
        'pendientes': pendientes,
        'en_curso': en_curso,
        'implementadas': implementadas,
        'continuas': continuas,
        'evitables': evitables,
        'monitorizables': monitorizables,
    }


def importar_planificacion_excel(archivo, empresa):
    from openpyxl import load_workbook
    from .models import ItemPlanificacion

    wb = load_workbook(archivo, read_only=True)
    ws = wb.active

    header_map = {}
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=False):
        for cell in row:
            if cell.value:
                val = str(cell.value).strip().lower()
                if 'ambito' in val or 'puesto' in val:
                    header_map['ambito'] = cell.column
                elif 'tipo de factor' in val:
                    header_map['tipo_factor'] = cell.column
                elif 'factor de riesgo' in val and 'detalle' not in val:
                    header_map['factor_riesgo'] = cell.column
                elif val == 'detalle':
                    header_map['detalle'] = cell.column
                elif 'riesgos' in val:
                    header_map['riesgos'] = cell.column
                elif 'pb' == val or 'probabilidad' in val:
                    header_map['pb'] = cell.column
                elif 'sv' == val or 'severidad' in val:
                    header_map['sv'] = cell.column
                elif 'gr' == val or 'grado' in val:
                    header_map['gr'] = cell.column
                elif 'medidas' in val and 'detalle' not in val:
                    header_map['medidas'] = cell.column
                elif 'detalle' in val and 'medida' in val:
                    header_map['detalle_medida'] = cell.column
                elif 'plazo' in val:
                    header_map['plazo'] = cell.column
                elif 'fecha' in val and 'objetivo' in val:
                    header_map['fecha_objetivo'] = cell.column
                elif 'responsable' in val:
                    header_map['responsable'] = cell.column
                elif 'coste' in val:
                    header_map['costes'] = cell.column
                elif 'estado' in val:
                    header_map['estado'] = cell.column
                elif 'origen' in val:
                    header_map['origen'] = cell.column

    start_row = 1
    for row in ws.iter_rows(min_row=1, max_row=5, max_col=1):
        for cell in row:
            if cell.value and ('ambito' in str(cell.value).lower() or 'puesto' in str(cell.value).lower()):
                start_row = cell.row + 1
                break

    def get_val(row, key, default=''):
        col = header_map.get(key)
        if col:
            val = ws.cell(row=row, column=col).value
            return str(val).strip() if val is not None else default
        return default

    importados = 0
    errores = []
    tipo_map = dict(ItemPlanificacion.TipoFactorRiesgo.choices)
    estado_map = dict(ItemPlanificacion.Estado.choices)
    riesgo_map = {v.lower(): k for k, v in ItemPlanificacion.RIESGO_CHOICES}

    for row_num in range(start_row, ws.max_row + 1):
        factor_riesgo = get_val(row_num, 'factor_riesgo')
        if not factor_riesgo:
            continue

        tipo_texto = get_val(row_num, 'tipo_factor', 'evitables').lower()
        tipo_valor = 'evitables'
        for k, v in tipo_map.items():
            if v.lower() == tipo_texto or k == tipo_texto:
                tipo_valor = k
                break

        estado_texto = get_val(row_num, 'estado', 'pendiente').lower()
        estado_valor = 'pendiente'
        for k, v in estado_map.items():
            if v.lower() == estado_texto or k == estado_texto:
                estado_valor = k
                break

        riesgos_texto = get_val(row_num, 'riesgos').lower()
        riesgo_valor = ''
        for texto, clave in riesgo_map.items():
            if texto in riesgos_texto or riesgos_texto in texto:
                riesgo_valor = clave
                break

        fecha_obj = None
        fecha_str = get_val(row_num, 'fecha_objetivo')
        if fecha_str and fecha_str != '-':
            from datetime import datetime
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
                try:
                    fecha_obj = datetime.strptime(fecha_str[:10], fmt).date()
                    break
                except ValueError:
                    continue

        plazo = get_val(row_num, 'plazo')
        if plazo == '-':
            plazo = ''

        try:
            ItemPlanificacion.objects.create(
                empresa=empresa,
                ambito_puesto=get_val(row_num, 'ambito'),
                tipo_factor_riesgo=tipo_valor,
                factor_riesgo=factor_riesgo,
                detalle=get_val(row_num, 'detalle'),
                riesgos=riesgo_valor,
                pb=get_val(row_num, 'pb', 'NV').upper()[:2],
                sv=get_val(row_num, 'sv', 'NV').upper()[:2],
                gr=get_val(row_num, 'gr', 'NV').upper()[:2],
                medidas_preventivas=get_val(row_num, 'medidas'),
                detalle_medida=get_val(row_num, 'detalle_medida'),
                plazo_limite=plazo,
                fecha_objetivo=fecha_obj,
                responsable=get_val(row_num, 'responsable'),
                costes=get_val(row_num, 'costes'),
                estado=estado_valor,
                origen=get_val(row_num, 'origen'),
            )
            importados += 1
        except Exception as e:
            errores.append(f'Fila {row_num}: {str(e)}')

    wb.close()
    return importados, errores
